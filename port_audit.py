#!/usr/bin/env python3
"""
port_audit.py  –  Cisco Switch Port Drop & Counter Auditor
===========================================================
Verified against C9500 (TwentyFiveGigE / HundredGigE / AppGigabitEthernet)
running IOS-XE with dual-stack interface name formats.

Usage:
    python port_audit.py --targets devices.txt
    python port_audit.py --targets devices.txt --agg 10.0.0.1
    python port_audit.py --targets devices.txt --no-jump
    python port_audit.py --targets devices.txt --out site_a_audit.xlsx
"""

import paramiko
import time
import re
import sys
import argparse
import logging
import warnings
import os
from datetime import datetime
from collections import OrderedDict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ============================================================================
# CONFIGURATION
# ============================================================================
AGG_SWITCH_IP = "192.168.0.251"

CREDENTIAL_SETS = [
    {"username": "admin",  "password": "admin",                "enable": "flounder"},
    {"username": "Admin",  "password": "/2/_HKX6YvCGMwzAdJp",  "enable": ""},
]

CMD_TIMEOUT  = 60
JUMP_TIMEOUT = 30

RETRY_COUNT = 2    # How many times to retry a failed connection (not counting the first attempt)
RETRY_DELAY = 5    # Seconds to wait between retries

CLEAR_WAIT  = 300  # Seconds to wait after 'clear counters' before collecting (default: 5 min)

# ============================================================================
# INTERFACE NAME NORMALIZATION
# ============================================================================
# IOS-XE uses mixed forms across commands:
#   show ip int brief  -> TwentyFiveGigE1/0/1  OR  Te1/1/1  OR  Ap1/0/1
#   show interfaces    -> TwentyFiveGigE1/0/1 is up ...
#   show int trunk     -> Twe1/0/1
#   show int count err -> Twe1/0/1 / Ap1/0/1 / Po1
#
# Strategy: map everything to a short lowercase key for cross-referencing,
# keep original display name from 'show ip int brief'.

_FULL_TO_SHORT = [
    ("twentyfivegige",       "twe"),
    ("tengigabitethernet",   "te"),
    ("gigabitethernet",      "gi"),
    ("fastethernet",         "fa"),
    ("hundredgige",          "hu"),
    ("appgigabitethernet",   "ap"),
    ("port-channel",         "po"),
    ("portchannel",          "po"),
    ("loopback",             "lo"),
    ("vlan",                 "vl"),
    ("tunnel",               "tu"),
]

def intf_key(name: str) -> str:
    """
    Canonical lowercase short key from any interface name form.
    TwentyFiveGigE1/0/1 -> twe1/0/1
    Twe1/0/1            -> twe1/0/1
    Port-channel1       -> po1
    Po1                 -> po1
    AppGigabitEthernet1/0/1 -> ap1/0/1
    Ap1/0/1             -> ap1/0/1
    """
    n = name.strip().lower().replace(" ", "")
    for full, short in _FULL_TO_SHORT:
        n = n.replace(full, short)
    return n

# ============================================================================
# COLUMN DEFINITIONS
# ============================================================================
COLUMNS = [
    "Interface",
    "Description",
    "Mode",
    "Speed",
    "Duplex",
    "Native VLAN",
    "Allowed VLANs",
    "In Rate (bps)",
    "In Rate (pps)",
    "Out Rate (bps)",
    "Out Rate (pps)",
    "Pkts In",
    "Bytes In",
    "No Buffer",
    "Input Errors",
    "CRC",
    "Frame",
    "Overruns",
    "Ignored",
    "Runts",
    "Giants",
    "Throttles",
    "Dribble",
    "Pkts Out",
    "Bytes Out",
    "Output Errors",
    "Collisions",
    "Intf Resets",
    "Unknown Proto Drops",
    "Out Buffer Failures",
    "Out Bufs Swapped",
    "Late Collision",
    "Deferred",
    "Lost Carrier",
    "No Carrier",
    # From show interfaces counters errors
    "Align Err",
    "FCS Err",
    "Xmit Err",
    "Rcv Err",
    "Undersize",
    "Out Discards",
    "Single Col",
    "Multi Col",
    "Excess Col",
    "Carri Sen",
    "Oversize",
    # From show interfaces transceiver (DOM)
    "SFP Type",
    "Temp (C)",
    "Voltage (V)",
    "Current (mA)",
    "Tx Power (dBm)",
    "Rx Power (dBm)",
    "SFP Alarms",      # human-readable alarm list e.g. "LOW-RX, HIGH-TEMP"
    "SFP Status",      # OK | WARN | ALARM | NOT PRESENT | N/A
    # From show interfaces transceiver detail (fiber SFP thresholds)
    "Rx Hi Alarm (dBm)",   # DOM threshold — high alarm
    "Rx Hi Warn (dBm)",    # DOM threshold — high warning
    "Rx Lo Warn (dBm)",    # DOM threshold — low warning  (most useful for fiber)
    "Rx Lo Alarm (dBm)",   # DOM threshold — low alarm
    "Tx Hi Alarm (dBm)",
    "Tx Hi Warn (dBm)",
    "Tx Lo Warn (dBm)",
    "Tx Lo Alarm (dBm)",
    "Rx Margin (dB)",      # Rx Power minus Rx Lo Warn — negative means below threshold
    "Fiber Rx Health",     # OK / NEAR WARN / BELOW WARN / BELOW ALARM / N/A
    # From show lldp neighbors detail
    "LLDP Neighbor",     # System name
    "LLDP Neighbor MAC", # Chassis MAC from LLDP
    "LLDP Neighbor IP",  # Management IP from LLDP
    # From show mac address-table (populated when no LLDP entry exists)
    "MAC Table Entries",  # All unique MACs seen on the port, semicolon-separated
    # Summary
    "Has Drops?",
]

DROP_COLUMNS = {
    "No Buffer", "Input Errors", "CRC", "Frame", "Overruns", "Ignored",
    "Runts", "Giants", "Throttles", "Dribble",
    "Output Errors", "Collisions",
    "Unknown Proto Drops", "Out Buffer Failures", "Out Bufs Swapped",
    "Late Collision", "Lost Carrier",
    "Align Err", "FCS Err", "Xmit Err", "Rcv Err",
    "Undersize", "Out Discards", "Oversize",
    # SFP alarm/warning triggers the drop flag too
    "SFP Alarms",
}

# ============================================================================
# LOGGING
# ============================================================================
class _ConsoleFmt(logging.Formatter):
    def format(self, record):
        ts = datetime.fromtimestamp(record.created).strftime('%H:%M:%S')
        return f"[{ts}] {record.getMessage()}"

def setup_logging():
    log_file = f"port_audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(_ConsoleFmt())
    root.addHandler(ch)
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    root.addHandler(fh)
    logging.getLogger("paramiko").setLevel(logging.WARNING)
    return logging.getLogger(__name__), log_file

logger = None

# ============================================================================
# LOGIN STATISTICS
# ============================================================================
class LoginStats:
    def __init__(self):
        self._log = []

    def record(self, ip, cred_idx, success, hostname="", error=""):
        self._log.append(dict(ip=ip, cred_idx=cred_idx, success=success,
                              hostname=hostname, error=error))

    def print_summary(self):
        total   = len(self._log)
        success = sum(1 for r in self._log if r["success"])
        failed  = total - success

        # Compute per-IP final outcome (last record for each IP wins)
        ip_final = {}
        for r in self._log:
            ip_final[r["ip"]] = r

        div = "=" * 72
        print(f"\n{div}")
        print("  LOGIN STATISTICS SUMMARY")
        print(div)
        print(f"  Total login attempts  : {total}  "
              f"(includes retries — {len(ip_final)} unique IPs)")
        print(f"  Successful attempts   : {success}")
        print(f"  Failed attempts       : {failed}")
        print(f"  IPs successfully done : {sum(1 for r in ip_final.values() if r['success'])}")
        print(f"  IPs permanently failed: {sum(1 for r in ip_final.values() if not r['success'])}")
        print("-" * 72)
        print(f"  {'IP':<22} {'Attempts':<10} {'Final':<10} Detail")
        print("-" * 72)
        # Group by IP to count attempts
        from collections import Counter
        ip_attempts = Counter(r["ip"] for r in self._log)
        for ip, final in ip_final.items():
            attempts = ip_attempts[ip]
            result   = "OK" if final["success"] else "FAIL"
            detail   = final["hostname"] if final["success"] else final["error"]
            retry_note = f" ({attempts} attempt{'s' if attempts > 1 else ''})"
            print(f"  {ip:<22} {retry_note:<10} {result:<10} {detail}")
        print(div + "\n")

stats = LoginStats()

# ============================================================================
# SSH UTILITIES
# ============================================================================
def _drain(shell):
    try:
        time.sleep(0.5)
        buf = ""
        while shell and not shell.closed and shell.recv_ready():
            buf += shell.recv(65535).decode("utf-8", "ignore")
        return buf
    except:
        return ""

def send_cmd(shell, cmd, timeout=CMD_TIMEOUT, silent=False):
    if not silent:
        logger.debug(f"CMD: {cmd}")
    try:
        if shell is None or shell.closed:
            raise OSError(f"Socket closed before sending: {cmd}")
        _drain(shell)
        shell.send(cmd + "\n")
        time.sleep(0.5)
        buf, end = "", time.time() + timeout
        while time.time() < end:
            if shell.closed:
                raise OSError(f"Socket closed while waiting for response to: {cmd}")
            if shell.recv_ready():
                chunk = shell.recv(65535).decode("utf-8", "ignore")
                buf += chunk
                if re.search(r"(?m)^[^\r\n>\s][^\r\n>]*[>#]\s?$", buf):
                    return buf
            time.sleep(0.2)
        return buf
    except OSError:
        raise   # let caller handle socket failures explicitly
    except Exception:
        return ""

def get_hostname(shell):
    """
    Reliably extract the hostname from the current shell session.

    Strategy:
      1. Tickle the prompt up to 3 times with \n and read back the prompt line.
         The prompt on Cisco IOS-XE is  HOSTNAME#  or  HOSTNAME>
      2. If prompt-parsing fails (shell still settling, ANSI garbage, long MOTD),
         fall back to 'show version | include hostname' which is unambiguous.
    """
    # ── Pass 1: read the prompt directly (fast path) ────────────────────────
    for _ in range(3):
        try:
            shell.send("\n"); time.sleep(0.8)
            buf = _drain(shell)
            for line in reversed([l.strip() for l in buf.splitlines() if l.strip()]):
                # Strip ANSI escape codes before matching
                clean = re.sub(r"\x1b\[[0-9;]*[A-Za-z]", "", line)
                if clean.endswith(("#", ">")):
                    name = re.sub(r"^[^\w]+", "", clean[:-1].strip())
                    # Sanity-check: must look like a hostname (≥2 chars, no spaces)
                    if name and len(name) >= 2 and " " not in name:
                        return name
        except Exception:
            pass

    # ── Pass 2: ask the device directly via show version ────────────────────
    try:
        shell.send("show version | include [Ss]ystem [Nn]ame\n")
        time.sleep(1.5)
        buf = _drain(shell)
        m = re.search(r"[Ss]ystem [Nn]ame\s*[:\-]?\s*(\S+)", buf)
        if m:
            return m.group(1).strip()
    except Exception:
        pass

    # ── Pass 3: parse 'show running-config | include hostname' ──────────────
    try:
        shell.send("show running-config | include ^hostname\n")
        time.sleep(1.5)
        buf = _drain(shell)
        m = re.search(r"^hostname\s+(\S+)", buf, re.M)
        if m:
            return m.group(1).strip()
    except Exception:
        pass

    return "Unknown"

# ============================================================================
# AGG / JUMP HOST
# ============================================================================
agg_client   = None
agg_shell    = None
agg_hostname = "Unknown"

def connect_to_agg(agg_ip):
    global agg_client, agg_shell, agg_hostname
    for attempt in range(1, RETRY_COUNT + 2):   # attempts = 1 + RETRY_COUNT
        if attempt > 1:
            logger.info(f"  [Retry {attempt - 1}/{RETRY_COUNT}] Reconnecting to aggregate {agg_ip} in {RETRY_DELAY}s ...")
            time.sleep(RETRY_DELAY)
        else:
            logger.info(f"Connecting to aggregate {agg_ip} ...")
        for i, cred in enumerate(CREDENTIAL_SETS):
            try:
                client = paramiko.SSHClient()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                client.connect(agg_ip, username=cred["username"],
                               password=cred["password"], timeout=20)
                client.get_transport().set_keepalive(20)
                shell = client.invoke_shell()
                time.sleep(1)
                banner = _drain(shell)
                if ">" in banner and "#" not in banner:
                    shell.send("enable\n"); time.sleep(0.5)
                    shell.send(cred.get("enable", cred["password"]) + "\n"); time.sleep(1)
                send_cmd(shell, "terminal length 0", silent=True)
                agg_hostname = get_hostname(shell)
                agg_client, agg_shell = client, shell
                logger.info(f"  Aggregate connected: {agg_hostname}")
                stats.record(agg_ip, i, True, hostname=agg_hostname)
                return True
            except Exception as e:
                stats.record(agg_ip, i, False, error=str(e))
        logger.warning(f"  All credentials failed on attempt {attempt} for {agg_ip}")
    logger.error(f"  Gave up connecting to aggregate {agg_ip} after {RETRY_COUNT + 1} attempts")
    return False

def jump_to_target(target_ip):
    global agg_shell
    for attempt in range(1, RETRY_COUNT + 2):
        if attempt > 1:
            logger.info(f"  [Retry {attempt - 1}/{RETRY_COUNT}] Jumping to {target_ip} in {RETRY_DELAY}s ...")
            time.sleep(RETRY_DELAY)
            return_to_agg()   # ensure we're back at a clean agg prompt before retrying
        for i, cred in enumerate(CREDENTIAL_SETS):
            try:
                agg_shell.send("\x03"); time.sleep(0.3); _drain(agg_shell)
                agg_shell.send(f"ssh -l {cred['username']} {target_ip}\n")
                buf, start = "", time.time()
                while (time.time() - start) < JUMP_TIMEOUT:
                    if agg_shell.recv_ready():
                        chunk = agg_shell.recv(65535).decode("utf-8", "ignore")
                        buf += chunk
                        if "password:" in chunk.lower():
                            agg_shell.send(cred["password"] + "\n")
                        if "yes/no" in chunk.lower():
                            agg_shell.send("yes\n")
                        if re.search(r"(?m)^[^\r\n>\s][^\r\n>]*[>#]\s?$", buf):
                            if ">" in buf and "#" not in buf:
                                agg_shell.send("enable\n"); time.sleep(0.5)
                                agg_shell.send(cred.get("enable", cred["password"]) + "\n")
                            send_cmd(agg_shell, "terminal length 0", silent=True)
                            hostname = get_hostname(agg_shell)
                            logger.info(f"  Jumped to {hostname} ({target_ip})"
                                        + (f" [attempt {attempt}]" if attempt > 1 else ""))
                            stats.record(target_ip, i, True, hostname=hostname)
                            return hostname
                    time.sleep(0.3)
            except Exception as e:
                stats.record(target_ip, i, False, error=str(e))
        logger.warning(f"  All credentials failed on attempt {attempt} for {target_ip}")
    stats.record(target_ip, len(CREDENTIAL_SETS) - 1, False,
                 error=f"Failed after {RETRY_COUNT + 1} attempts")
    logger.error(f"  Gave up jumping to {target_ip} after {RETRY_COUNT + 1} attempts")
    return False

def connect_direct(target_ip):
    for attempt in range(1, RETRY_COUNT + 2):
        if attempt > 1:
            logger.info(f"  [Retry {attempt - 1}/{RETRY_COUNT}] Connecting to {target_ip} in {RETRY_DELAY}s ...")
            time.sleep(RETRY_DELAY)
        for i, cred in enumerate(CREDENTIAL_SETS):
            try:
                client = paramiko.SSHClient()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                client.connect(target_ip, username=cred["username"],
                               password=cred["password"], timeout=20)
                client.get_transport().set_keepalive(20)
                shell = client.invoke_shell()
                time.sleep(1)
                banner = _drain(shell)
                if ">" in banner and "#" not in banner:
                    shell.send("enable\n"); time.sleep(0.5)
                    shell.send(cred.get("enable", cred["password"]) + "\n"); time.sleep(1)
                send_cmd(shell, "terminal length 0", silent=True)
                hostname = get_hostname(shell)
                logger.info(f"  Connected to {hostname} ({target_ip})"
                            + (f" [attempt {attempt}]" if attempt > 1 else ""))
                stats.record(target_ip, i, True, hostname=hostname)
                return client, shell, hostname
            except Exception as e:
                stats.record(target_ip, i, False, error=str(e))
        logger.warning(f"  All credentials failed on attempt {attempt} for {target_ip}")
    logger.error(f"  Gave up connecting to {target_ip} after {RETRY_COUNT + 1} attempts")
    return None, None, None

def return_to_agg():
    global agg_shell
    # If the socket is already dead, skip straight to full reconnect
    try:
        if agg_shell is None or agg_shell.closed:
            raise OSError("shell is closed")
        agg_shell.send("\x03")
        for _ in range(3):
            agg_shell.send("exit\n")
            time.sleep(1)
            buf = _drain(agg_shell)
            if agg_hostname.lower() in buf.lower():
                return True
    except OSError:
        logger.warning("  Agg shell socket dropped — reconnecting ...")
    except Exception as e:
        logger.warning(f"  return_to_agg unexpected error ({e}) — reconnecting ...")
    return connect_to_agg(AGG_SWITCH_IP)

# ============================================================================
# COMMAND COLLECTION
# ============================================================================
def clear_counters(shell):
    """
    Issues 'clear counters' and confirms the [confirm] prompt.
    Raises OSError if the socket drops.
    """
    logger.info("  [*] clear counters (sending confirmation) ...")
    # Send the command — IOS will reply with "[confirm]" and wait
    send_cmd(shell, "clear counters", timeout=15, silent=True)
    # Send a blank line to confirm; IOS accepts Enter as confirmation
    shell.send("\n")
    time.sleep(2)
    _drain(shell)   # consume the confirmation echo


def run_commands(shell):
    # ── Step 0: clear counters then wait ─────────────────────────────────────
    clear_counters(shell)

    logger.info(f"  Waiting {CLEAR_WAIT}s for traffic to accumulate ...")
    wait_end = time.time() + CLEAR_WAIT
    while time.time() < wait_end:
        remaining = int(wait_end - time.time())
        if remaining % 60 == 0 and remaining > 0:
            logger.info(f"    ... {remaining // 60}m remaining ...")
        # Keep-alive: send a harmless newline every 30s so the SSH session
        # does not idle-timeout while we wait
        if int(time.time()) % 30 == 0:
            try:
                shell.send("\n")
                _drain(shell)
            except OSError:
                raise
        time.sleep(1)

    # ── Step 1-6: collect ─────────────────────────────────────────────────────
    logger.info("  [1/6] show ip interface brief")
    ip_brief    = send_cmd(shell, "show ip interface brief")
    logger.info("  [2/6] show interfaces")
    intf_full   = send_cmd(shell, "show interfaces", timeout=90)
    logger.info("  [3/6] show interfaces status")
    intf_status = send_cmd(shell, "show interfaces status")
    logger.info("  [4/6] show interfaces trunk")
    intf_trunk  = send_cmd(shell, "show interfaces trunk")
    logger.info("  [5/6] show interfaces counters errors")
    intf_errs   = send_cmd(shell, "show interfaces counters errors")
    logger.info("  [6/6] show interfaces transceiver")
    transceiver = send_cmd(shell, "show interfaces transceiver")
    logger.info("  [7/9] show lldp neighbors detail")
    lldp        = send_cmd(shell, "show lldp neighbors detail", timeout=60)
    logger.info("  [8/9] show mac address-table dynamic")
    mac_table   = send_cmd(shell, "show mac address-table dynamic", timeout=60)
    logger.info("  [9/10] show interfaces transceiver detail")
    xcvr_detail  = send_cmd(shell, "show interfaces transceiver detail", timeout=60)
    logger.info("  [10/10] show etherchannel summary")
    etherchannel = send_cmd(shell, "show etherchannel summary")
    return dict(ip_brief=ip_brief, intf_full=intf_full,
                intf_status=intf_status, intf_trunk=intf_trunk,
                intf_errs=intf_errs, transceiver=transceiver,
                lldp=lldp, mac_table=mac_table, xcvr_detail=xcvr_detail,
                etherchannel=etherchannel)

# ============================================================================
# PARSERS
# ============================================================================
def _int(val):
    try:
        return int(str(val).replace(",", "").strip())
    except:
        return 0


# Regex that matches every character illegal in OOXML cell strings
_ILLEGAL_XML = re.compile(
    u"[\u0000-\u0008\u000b\u000c\u000e-\u001f"
    u"\u007f\u0080-\u009f\ufffe\uffff]"
)
# ANSI escape sequences (ESC [ ... m and similar)
_ANSI_ESC = re.compile(r"\x1b\[[0-9;]*[A-Za-z]|\x1b[()=><]")

def _safe_val(val):
    """
    Sanitise a value before writing it to an openpyxl cell.
    - Strips ANSI escape codes and illegal XML characters from strings.
    - Normalises \r\n / \r to a single space so text stays on one line.
    - Converts None / float NaN / float Inf to empty string.
    - Leaves int / float (finite) unchanged.
    """
    import math
    if val is None:
        return ""
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return ""
        return val
    if isinstance(val, int):
        return val
    s = str(val)
    s = _ANSI_ESC.sub("", s)               # strip ANSI codes
    s = s.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    s = _ILLEGAL_XML.sub("", s)            # strip illegal XML chars
    s = s.strip()
    return s


def parse_up_interfaces(ip_brief_out):
    """
    Returns {intf_key: display_name} for every up/up physical interface.
    Excludes Vlan, Loopback, Tunnel, Null, Async.
    """
    result = {}
    SKIP = ("vlan", "loop", "null", "tunnel", "async")
    for line in ip_brief_out.splitlines():
        parts = line.split()
        if len(parts) < 6:
            continue
        name = parts[0]
        if parts[4].lower() != "up" or parts[5].lower() != "up":
            continue
        if any(name.lower().startswith(s) for s in SKIP):
            continue
        result[intf_key(name)] = name
    return result


def parse_intf_status(intf_status_out):
    """
    Parses 'show interfaces status'.
    C9500 format:
      Port     Name       Status      Vlan  Duplex  Speed   Type
      Twe1/0/1 UPLINK-SW  connected   trunk a-full  a-25G   SFP-25GBase-SR

    Returns {intf_key: {description, vlan, duplex, speed, port_type}}
    """
    STATUS_WORDS = {"connected", "notconnect", "disabled",
                    "err-disabled", "inactive", "sfpabsent"}
    result, header_found = {}, False

    for line in intf_status_out.splitlines():
        if re.match(r"^\s*Port\s+", line, re.I):
            header_found = True
            continue
        if not header_found or not line.strip() or line.strip().startswith("-"):
            continue
        parts = line.split()
        if len(parts) < 2:
            continue
        port = parts[0]
        st_idx = next((i for i, p in enumerate(parts)
                       if p.lower() in STATUS_WORDS), None)
        if st_idx is None:
            continue
        desc     = " ".join(parts[1:st_idx]) if st_idx > 1 else ""
        trailing = parts[st_idx + 1:]
        result[intf_key(port)] = dict(
            description=desc,
            vlan    =trailing[0] if len(trailing) > 0 else "",
            duplex  =trailing[1] if len(trailing) > 1 else "",
            speed   =trailing[2] if len(trailing) > 2 else "",
            port_type=" ".join(trailing[3:]) if len(trailing) > 3 else "",
        )
    return result


def parse_trunk_ports(intf_trunk_out):
    """
    Parses 'show interfaces trunk'.

    Output has FOUR sequential sections:
      Section 1 header:  Port  Mode  Encapsulation  Status  Native vlan
      Section 2 header:  Port  Vlans allowed on trunk
      Section 3 header:  Port  Vlans allowed and active in management domain
      Section 4 header:  Port  Vlans in spanning tree forwarding state and not pruned

    Returns {intf_key: {native_vlan, allowed_vlans, active_vlans}}
    """
    result  = {}
    section = None

    for line in intf_trunk_out.splitlines():
        s = line.strip()
        if not s or s.startswith("-"):
            continue

        # Section header detection (order matters)
        if re.match(r"^Port\s+Mode\s+Encapsulation", s, re.I):
            section = "mode";    continue
        if re.match(r"^Port\s+Vlans allowed on trunk", s, re.I):
            section = "allowed"; continue
        if re.match(r"^Port\s+Vlans allowed and active", s, re.I):
            section = "active";  continue
        if re.match(r"^Port\s+Vlans in spanning", s, re.I):
            section = "stp";     continue
        # Generic "Port ..." header we don't recognise -> skip
        if s.lower().startswith("port") and section is None:
            continue

        parts = s.split()
        if len(parts) < 2:
            continue
        key = intf_key(parts[0])

        if section == "mode":
            # parts: Port Mode Encapsulation Status Native-vlan
            if key not in result:
                result[key] = {}
            result[key]["native_vlan"] = parts[4] if len(parts) >= 5 else ""

        elif section == "allowed":
            if key not in result:
                result[key] = {}
            result[key]["allowed_vlans"] = parts[1]

        elif section == "active":
            if key not in result:
                result[key] = {}
            result[key]["active_vlans"] = parts[1]

    return result


def parse_intf_full(intf_full_out, up_keys):
    """
    Parses 'show interfaces'. Only processes interfaces whose key is in up_keys.
    Returns {intf_key: {counter_name: value}}
    """
    blocks = re.split(r"\n(?=[A-Za-z])", intf_full_out)
    result = {}

    for block in blocks:
        lines = block.splitlines()
        if not lines:
            continue
        m = re.match(r"^(\S+)\s+is\s+(up|down|administratively down)", lines[0], re.I)
        if not m:
            continue
        key = intf_key(m.group(1))
        if key not in up_keys:
            continue

        d = {}

        # Description
        dm = re.search(r"Description:\s*(.*)", block)
        d["Description"] = dm.group(1).strip() if dm else ""

        # Speed / Duplex from hardware capability line
        hw = re.search(
            r"((?:Full|Half|Auto|a-full|a-half)[- ]duplex),\s*([\d.]+\s*\S+b/s)",
            block, re.I)
        if hw:
            d["Duplex"] = hw.group(1)
            d["Speed"]  = hw.group(2)

        # 5-minute rates — two occurrences (input then output)
        rates = re.findall(r"(\d[\d,]*)\s+bits/sec,\s+(\d[\d,]*)\s+packets/sec", block)
        if len(rates) >= 1:
            d["In Rate (bps)"]  = _int(rates[0][0])
            d["In Rate (pps)"]  = _int(rates[0][1])
        if len(rates) >= 2:
            d["Out Rate (bps)"] = _int(rates[1][0])
            d["Out Rate (pps)"] = _int(rates[1][1])

        # Input packets / bytes / no-buffer
        m2 = re.search(
            r"(\d[\d,]*)\s+packets input,\s+(\d[\d,]*)\s+bytes,\s+(\d[\d,]*)\s+no buffer",
            block)
        if m2:
            d["Pkts In"]   = _int(m2.group(1))
            d["Bytes In"]  = _int(m2.group(2))
            d["No Buffer"] = _int(m2.group(3))

        # Output packets / bytes
        m3 = re.search(r"(\d[\d,]*)\s+packets output,\s+(\d[\d,]*)\s+bytes", block)
        if m3:
            d["Pkts Out"]  = _int(m3.group(1))
            d["Bytes Out"] = _int(m3.group(2))

        # Input errors
        m4 = re.search(
            r"(\d[\d,]*)\s+input errors,\s+(\d[\d,]*)\s+CRC,\s+"
            r"(\d[\d,]*)\s+frame,\s+(\d[\d,]*)\s+overrun,\s+(\d[\d,]*)\s+ignored",
            block)
        if m4:
            d["Input Errors"] = _int(m4.group(1))
            d["CRC"]          = _int(m4.group(2))
            d["Frame"]        = _int(m4.group(3))
            d["Overruns"]     = _int(m4.group(4))
            d["Ignored"]      = _int(m4.group(5))

        # Runts / giants / throttles
        m5 = re.search(
            r"(\d[\d,]*)\s+runts,\s+(\d[\d,]*)\s+giants,\s+(\d[\d,]*)\s+throttles",
            block)
        if m5:
            d["Runts"]     = _int(m5.group(1))
            d["Giants"]    = _int(m5.group(2))
            d["Throttles"] = _int(m5.group(3))

        # Dribble
        m6 = re.search(r"(\d[\d,]*)\s+input packets with dribble", block)
        if m6:
            d["Dribble"] = _int(m6.group(1))

        # Output errors / collisions / resets
        m7 = re.search(
            r"(\d[\d,]*)\s+output errors,\s+(\d[\d,]*)\s+collisions,\s+"
            r"(\d[\d,]*)\s+interface resets",
            block)
        if m7:
            d["Output Errors"] = _int(m7.group(1))
            d["Collisions"]    = _int(m7.group(2))
            d["Intf Resets"]   = _int(m7.group(3))

        # Unknown protocol drops
        m8 = re.search(r"(\d[\d,]*)\s+unknown protocol drops", block)
        if m8:
            d["Unknown Proto Drops"] = _int(m8.group(1))

        # Output buffer failures / swapped
        m9 = re.search(
            r"(\d[\d,]*)\s+output buffer failures,\s+(\d[\d,]*)\s+output buffers swapped out",
            block)
        if m9:
            d["Out Buffer Failures"] = _int(m9.group(1))
            d["Out Bufs Swapped"]    = _int(m9.group(2))

        # Babbles / late collision / deferred
        m10 = re.search(
            r"(\d[\d,]*)\s+babbles,\s+(\d[\d,]*)\s+late collision,\s+(\d[\d,]*)\s+deferred",
            block)
        if m10:
            d["Late Collision"] = _int(m10.group(2))
            d["Deferred"]       = _int(m10.group(3))

        # Lost / no carrier
        m11 = re.search(r"(\d[\d,]*)\s+lost carrier,\s+(\d[\d,]*)\s+no carrier", block)
        if m11:
            d["Lost Carrier"] = _int(m11.group(1))
            d["No Carrier"]   = _int(m11.group(2))

        result[key] = d

    return result


def parse_intf_counters_errors(intf_errs_out):
    """
    Parses 'show interfaces counters errors'.

    C9500 produces THREE sub-tables:

    Sub-table 1 (header contains 'Align-Err'):
      Port  Align-Err  FCS-Err  Xmit-Err  Rcv-Err  UnderSize  OutDiscards

    Sub-table 2 (header contains 'Single-Col'):
      Port  Single-Col  Multi-Col  Late-Col  Excess-Col  Carri-Sen  Runts

    Sub-table 3 (header is just 'Port  OverSize'):
      Port  OverSize

    Returns {intf_key: {column_name: value}}
    """
    result  = {}
    section = None

    HDR1 = re.compile(r"Align-Err",  re.I)
    HDR2 = re.compile(r"Single-Col", re.I)
    HDR3 = re.compile(r"^\s*Port\s+OverSize\s*$", re.I)

    for line in intf_errs_out.splitlines():
        s = line.strip()
        if not s:
            continue
        if HDR1.search(s):
            section = 1; continue
        if HDR2.search(s):
            section = 2; continue
        if HDR3.match(s):
            section = 3; continue
        if re.match(r"^-+$", s):
            continue

        parts = s.split()
        if len(parts) < 2:
            continue
        key = intf_key(parts[0])
        if key not in result:
            result[key] = {}

        if section == 1 and len(parts) >= 7:
            result[key]["Align Err"]    = _int(parts[1])
            result[key]["FCS Err"]      = _int(parts[2])
            result[key]["Xmit Err"]     = _int(parts[3])
            result[key]["Rcv Err"]      = _int(parts[4])
            result[key]["Undersize"]    = _int(parts[5])
            result[key]["Out Discards"] = _int(parts[6])

        elif section == 2 and len(parts) >= 7:
            result[key]["Single Col"] = _int(parts[1])
            result[key]["Multi Col"]  = _int(parts[2])
            # parts[3] = Late-Col (already captured from show interfaces)
            result[key]["Excess Col"] = _int(parts[4])
            result[key]["Carri Sen"]  = _int(parts[5])
            # parts[6] = Runts (already captured from show interfaces)

        elif section == 3 and len(parts) >= 2:
            result[key]["Oversize"] = _int(parts[1])

    return result


def parse_transceiver(transceiver_out):
    """
    Parses 'show interfaces transceiver'.

    C9500 / IOS-XE format — two possible layouts:

    Layout A  (standard DOM table):
    ─────────────────────────────────────────────────────────────────────────
                                               Optical   Optical
               Temperature  Voltage  Current   Tx Power  Rx Power
    Port       (Celsius)    (Volts)  (mA)      (dBm)     (dBm)
    ---------  -----------  -------  --------  --------  --------
    Twe1/0/1    35.1        3.28     6.80      -2.51     -3.14
    Twe1/0/2    35.0        3.28     6.79      -2.50     N/A++
    ─────────────────────────────────────────────────────────────────────────

    Layout B  (block per interface, seen on some IOS-XE versions):
    ─────────────────────────────────────────────────────────────────────────
    Twe1/0/1
      Temperature    : 35.1 C
      Voltage        : 3.28 V
      Current        : 6.80 mA
      Optical Tx Pwr : -2.51 dBm
      Optical Rx Pwr : -3.14 dBm
    ─────────────────────────────────────────────────────────────────────────

    Alarm/warning markers (appended to a value when threshold crossed):
      ++  high alarm
      +   high warning
      -   low warning
      --  low alarm

    "Not Present" / "hardware is not present" lines indicate missing SFP.
    "N/A" values mean DOM not supported by that SFP.

    Returns:
      { intf_key: {
          "Temp (C)": float|str, "Voltage (V)": float|str,
          "Current (mA)": float|str, "Tx Power (dBm)": float|str,
          "Rx Power (dBm)": float|str,
          "SFP Alarms": str,   # comma-separated list, "" if clean
          "SFP Status": str,   # OK | WARN | ALARM | NOT PRESENT | N/A
      } }
    """

    # Alarm marker → label
    ALARM_MARKERS = {
        "++": "HIGH-ALARM",
        "--": "LOW-ALARM",
        "+":  "HIGH-WARN",
        "-":  "LOW-WARN",
    }

    # Column index → field name, used for Layout A to track which column
    # an alarm marker belongs to.
    COL_FIELDS = ["Temp (C)", "Voltage (V)", "Current (mA)",
                  "Tx Power (dBm)", "Rx Power (dBm)"]

    result = {}

    def _sfp_float(tok):
        """Strip alarm marker suffix and convert to float. Returns (float|str, marker)."""
        tok = tok.strip()
        marker = ""
        for mk in ("++", "--", "+", "-"):   # order: longest first
            if tok.endswith(mk):
                marker = mk
                tok = tok[:-len(mk)]
                break
        try:
            return float(tok), marker
        except ValueError:
            return tok, marker  # "N/A", "---", etc.

    def _severity(marker):
        if "ALARM" in marker:
            return 2
        if "WARN" in marker:
            return 1
        return 0

    def _status_from_alarms(alarm_list):
        if not alarm_list:
            return "OK"
        if any("ALARM" in a for a in alarm_list):
            return "ALARM"
        return "WARN"

    # ── Detect layout ────────────────────────────────────────────────────────
    # Layout A: has a separator line "----  ----  ----"
    # Layout B: has "Temperature    :" style lines
    is_layout_b = bool(re.search(r"Temperature\s*:", transceiver_out, re.I))

    if is_layout_b:
        # ── Layout B: per-interface blocks ───────────────────────────────────
        # Split on lines that look like an interface name at column 0
        blocks = re.split(r"\n(?=[A-Za-z])", transceiver_out)
        for block in blocks:
            lines = block.splitlines()
            if not lines:
                continue
            first = lines[0].strip()
            # Interface name line?
            if not re.match(r"^[A-Za-z][\w/.-]+$", first):
                continue
            key = intf_key(first)
            d = {"SFP Alarms": "", "SFP Status": "N/A"}
            alarms = []

            if re.search(r"not present|not installed|hardware is not present",
                         block, re.I):
                d["SFP Status"] = "NOT PRESENT"
                result[key] = d
                continue

            field_map = {
                r"temperature"    : "Temp (C)",
                r"voltage"        : "Voltage (V)",
                r"current"        : "Current (mA)",
                r"optical tx"     : "Tx Power (dBm)",
                r"optical rx"     : "Rx Power (dBm)",
            }
            for line in lines[1:]:
                for pat, field in field_map.items():
                    if re.search(pat, line, re.I):
                        val_m = re.search(r":\s*([\S]+)", line)
                        if val_m:
                            val, mk = _sfp_float(val_m.group(1))
                            d[field] = val
                            if mk:
                                alarms.append(
                                    f"{field.split('(')[0].strip()}: "
                                    f"{ALARM_MARKERS.get(mk, mk)}")

            d["SFP Alarms"] = ", ".join(alarms)
            d["SFP Status"] = _status_from_alarms(alarms)
            result[key] = d

    else:
        # ── Layout A: flat table ──────────────────────────────────────────────
        in_table = False
        for line in transceiver_out.splitlines():
            # Separator line marks start of data rows
            if re.match(r"^-{5}", line.strip()):
                in_table = True
                continue
            if not in_table:
                continue
            stripped = line.strip()
            if not stripped:
                continue

            parts = stripped.split()
            if len(parts) < 2:
                continue

            # First token must look like an interface name
            if not re.match(r"^[A-Za-z][\w/.-]+$", parts[0]):
                continue

            key = intf_key(parts[0])
            values = parts[1:]   # up to 5 DOM values

            d = {"SFP Alarms": "", "SFP Status": "N/A"}
            alarms = []

            # "Not Present" rows have no numeric values
            if any(p.lower() in ("not", "present") for p in values):
                d["SFP Status"] = "NOT PRESENT"
                result[key] = d
                continue

            for i, tok in enumerate(values):
                if i >= len(COL_FIELDS):
                    break
                field = COL_FIELDS[i]
                val, mk = _sfp_float(tok)
                d[field] = val
                if mk:
                    alarms.append(
                        f"{field.split('(')[0].strip()}: "
                        f"{ALARM_MARKERS.get(mk, mk)}")

            d["SFP Alarms"] = ", ".join(alarms)
            d["SFP Status"] = _status_from_alarms(alarms)
            result[key] = d

    return result


def parse_lldp(lldp_out):
    """
    Parses 'show lldp neighbors detail'.

    Each neighbor block starts with a dashed separator and contains:
      Local Intf: Twe1/0/1
      System Name: SWITCH-ACC-01
      Chassis id: 0011.2233.aabb   (or in colon format 00:11:22:33:aa:bb)
      Management Addresses:
          IP: 10.1.1.10
        (or)
          IPv4: 10.1.1.10

    One interface can have multiple LLDP neighbors (rare but possible on hubs).
    We store the first/primary neighbor per local interface.

    Returns:
      { intf_key: { "hostname": str, "mac": str, "ip": str } }
    """
    result  = {}
    blocks  = re.split(r"-{10,}", lldp_out)

    for block in blocks:
        if not block.strip():
            continue

        # Local interface
        local_m = re.search(r"Local Intf(?:ace)?[:\s]+(\S+)", block, re.I)
        if not local_m:
            continue
        key = intf_key(local_m.group(1))

        # System name (hostname)
        sys_m = re.search(r"System Name[:\s]+(\S+)", block, re.I)
        hostname = sys_m.group(1).strip() if sys_m else ""

        # Chassis ID — LLDP uses MAC as chassis ID in most Cisco deployments
        # Formats seen: aabb.ccdd.eeff  or  aa:bb:cc:dd:ee:ff  or  aa-bb-cc-dd-ee-ff
        chassis_m = re.search(r"Chassis(?:[\s]+id)?[:\s]+([0-9a-fA-F]{2}[:\-\.]?[0-9a-fA-F]{2}"
                               r"[:\-\.]?[0-9a-fA-F]{2}[:\-\.]?[0-9a-fA-F]{2}"
                               r"[:\-\.]?[0-9a-fA-F]{2}[:\-\.]?[0-9a-fA-F]{2})", block, re.I)
        mac = ""
        if chassis_m:
            raw_mac = chassis_m.group(1)
            # Normalise to Cisco dotted-quad format xxxx.xxxx.xxxx
            clean = re.sub(r"[:\-\.]", "", raw_mac).lower()
            if len(clean) == 12:
                mac = ":".join(clean[i:i+2] for i in range(0, 12, 2)).upper()

        # Management / IP address — try both IPv4 and plain IP labels
        ip = ""
        ip_m = re.search(r"(?:Management\s+Address(?:es)?[^:]*:\s*)?(?:IP(?:v4)?)[:\s]+(\d{1,3}(?:\.\d{1,3}){3})",
                          block, re.I | re.DOTALL)
        if ip_m:
            ip = ip_m.group(1).strip()

        # Only write if we got at least something useful
        if hostname or mac or ip:
            # If multiple neighbors on same port, append rather than overwrite
            if key in result:
                existing = result[key]
                result[key] = {
                    "hostname": "; ".join(filter(None, [existing["hostname"], hostname])),
                    "mac":      "; ".join(filter(None, [existing["mac"],      mac])),
                    "ip":       "; ".join(filter(None, [existing["ip"],       ip])),
                }
            else:
                result[key] = {"hostname": hostname, "mac": mac, "ip": ip}

    return result


def parse_mac_table(mac_table_out):
    """
    Parses 'show mac address-table dynamic'.

    C9500 IOS-XE output format:
          Mac Address Table
    -------------------------------------------
    Vlan    Mac Address       Type        Ports
    ----    -----------       --------    -----
     100    001a.2b3c.4d5e   DYNAMIC     Twe1/0/3
     800    001a.2b3c.4d5f   DYNAMIC     Po1
     All    0000.0000.0000   STATIC      CPU        <- skipped by port filter

    Robustness notes:
    - Uses a regex per line rather than positional split so column padding
      and "All" VLAN values don't cause silent misses.
    - MAC is found by pattern anywhere in the line, port is the last token.
    - CPU / Router / Supervisor / Drop ports are excluded.
    - Works with both Cisco dotted (xxxx.xxxx.xxxx) and colon/dash formats.

    Returns:
      { intf_key: [ "AA:BB:CC:DD:EE:FF", ... ] }
    """
    result = {}

    # Ports we never want to record MACs for
    SKIP_PORTS = {"cpu", "router", "supervisor", "drop", "vlan", "igmp"}

    # Match a line that contains a MAC address — Cisco dotted or colon/dash
    MAC_PAT = re.compile(
        r"([0-9a-f]{4}\.[0-9a-f]{4}\.[0-9a-f]{4}"   # Cisco dotted
        r"|[0-9a-f]{2}[:\-][0-9a-f]{2}[:\-][0-9a-f]{2}"
        r"[:\-][0-9a-f]{2}[:\-][0-9a-f]{2}[:\-][0-9a-f]{2})",  # colon/dash
        re.I
    )

    for line in mac_table_out.splitlines():
        line = line.strip()
        if not line:
            continue

        # Must contain a MAC
        mac_m = MAC_PAT.search(line)
        if not mac_m:
            continue

        # Port is the last whitespace-separated token on the line
        parts = line.split()
        port  = parts[-1]

        # Skip internal/CPU ports
        if any(port.lower().startswith(s) for s in SKIP_PORTS):
            continue

        # Port must start with a letter and be at least 3 chars long.
        # This filters stray summary lines like "42" or "Total" while allowing
        # Po1, Gi0/1, Twe1/0/1, Ap1/0/1, Te1/1/1 etc.
        if not re.match(r"^[A-Za-z]{2}", port) or len(port) < 3:
            continue

        # Normalise MAC to XX:XX:XX:XX:XX:XX uppercase
        raw_mac = mac_m.group(1)
        clean   = re.sub(r"[:.\-]", "", raw_mac).lower()
        if len(clean) != 12 or not re.fullmatch(r"[0-9a-f]{12}", clean):
            continue
        mac_fmt = ":".join(clean[i:i+2] for i in range(0, 12, 2)).upper()

        key = intf_key(port)
        if key not in result:
            result[key] = []
        if mac_fmt not in result[key]:
            result[key].append(mac_fmt)

    return result


def parse_transceiver_detail(xcvr_detail_out):
    """
    Parses 'show interfaces transceiver detail'.

    Extracts per-interface threshold values for Rx and Tx optical power.
    Only produces entries for fiber SFP interfaces that support DOM.

    C9500 IOS-XE — two column orderings are seen in the wild:

    Layout A  (Current first):
    ─────────────────────────────────────────────────────────────────────────
    TwentyFiveGigE1/0/1
                              Current    High Alarm  High Warn   Low Warn  Low Alarm
      Temperature (Celsius) :  35.1        85.0        80.0        -5.0      -10.0
      Optical Tx Power (dBm):  -2.51        1.9         0.9        -8.2       -9.2
      Optical Rx Power (dBm):  -3.14        2.0        -1.0       -14.0      -15.0
    ─────────────────────────────────────────────────────────────────────────

    Layout B  (Thresholds first, current last):
    ─────────────────────────────────────────────────────────────────────────
    TwentyFiveGigE1/0/1
                               High Alarm  High Warn  Low Warn   Low Alarm
      Temperature (Celsius) :  85.0        80.0       -5.0       -10.0       35.1
      Optical Tx Power (dBm):   1.9         0.9        -8.2       -9.2       -2.51
      Optical Rx Power (dBm):   2.0        -1.0       -14.0      -15.0       -3.14
    ─────────────────────────────────────────────────────────────────────────

    Returns:
      { intf_key: {
          "Rx Hi Alarm (dBm)": float, "Rx Hi Warn (dBm)":  float,
          "Rx Lo Warn (dBm)":  float, "Rx Lo Alarm (dBm)": float,
          "Tx Hi Alarm (dBm)": float, "Tx Hi Warn (dBm)":  float,
          "Tx Lo Warn (dBm)":  float, "Tx Lo Alarm (dBm)": float,
      } }
    Only interfaces with valid optical readings are returned.
    """
    result  = {}

    # Split into per-interface blocks — interface name at column 0
    blocks = re.split(r"\n(?=[A-Za-z])", xcvr_detail_out)

    # Pattern to extract all numeric tokens (including negative floats) from a line
    NUM_PAT  = re.compile(r"-?\d+\.\d+|-?\d+")

    # Patterns to identify Rx / Tx power lines
    RX_PAT   = re.compile(r"optical\s+rx|rx\s+power", re.I)
    TX_PAT   = re.compile(r"optical\s+tx|tx\s+power", re.I)

    # Column header pattern — detects which layout we are in
    # Layout A has "Current" before "High Alarm"
    # Layout B has "High Alarm" without "Current" preceding it
    LAYOUT_A = re.compile(r"current.*high\s+alarm", re.I)
    LAYOUT_B = re.compile(r"high\s+alarm.*high\s+warn", re.I)

    def _safe_float(s):
        try:
            return float(s)
        except (TypeError, ValueError):
            return None

    for block in blocks:
        lines = block.splitlines()
        if not lines:
            continue

        # First non-empty line must be an interface name
        first = lines[0].strip()
        if not re.match(r"^[A-Za-z][\w/.-]+$", first):
            continue
        key = intf_key(first)

        # Skip if "not present" or no DOM support
        if re.search(r"not present|not installed|dom not supported|"
                     r"transceiver is not present|n/a", block, re.I):
            continue

        # Detect column layout from the header line inside this block
        layout = "B"   # default
        col_order_rx  = ["hi_alarm", "hi_warn", "lo_warn", "lo_alarm"]   # Layout B
        col_order_tx  = ["hi_alarm", "hi_warn", "lo_warn", "lo_alarm"]

        for line in lines:
            if LAYOUT_A.search(line):
                layout = "A"
                # Layout A: Current  Hi-Alarm  Hi-Warn  Lo-Warn  Lo-Alarm
                col_order_rx = ["current", "hi_alarm", "hi_warn", "lo_warn", "lo_alarm"]
                col_order_tx = ["current", "hi_alarm", "hi_warn", "lo_warn", "lo_alarm"]
                break
            if LAYOUT_B.search(line):
                layout = "B"
                # Layout B: Hi-Alarm  Hi-Warn  Lo-Warn  Lo-Alarm  [Current at end]
                col_order_rx = ["hi_alarm", "hi_warn", "lo_warn", "lo_alarm"]
                col_order_tx = ["hi_alarm", "hi_warn", "lo_warn", "lo_alarm"]
                break

        d = {}

        for line in lines[1:]:
            nums = NUM_PAT.findall(line)
            if not nums:
                continue

            if RX_PAT.search(line):
                for idx, field in enumerate(col_order_rx):
                    if idx < len(nums):
                        v = _safe_float(nums[idx])
                        if v is not None and field != "current":
                            key_name = {
                                "hi_alarm": "Rx Hi Alarm (dBm)",
                                "hi_warn":  "Rx Hi Warn (dBm)",
                                "lo_warn":  "Rx Lo Warn (dBm)",
                                "lo_alarm": "Rx Lo Alarm (dBm)",
                            }[field]
                            d[key_name] = v

            elif TX_PAT.search(line):
                for idx, field in enumerate(col_order_tx):
                    if idx < len(nums):
                        v = _safe_float(nums[idx])
                        if v is not None and field != "current":
                            key_name = {
                                "hi_alarm": "Tx Hi Alarm (dBm)",
                                "hi_warn":  "Tx Hi Warn (dBm)",
                                "lo_warn":  "Tx Lo Warn (dBm)",
                                "lo_alarm": "Tx Lo Alarm (dBm)",
                            }[field]
                            d[key_name] = v

        # Only store if we got at least Rx thresholds
        if "Rx Lo Warn (dBm)" in d or "Rx Lo Alarm (dBm)" in d:
            result[key] = d

    return result


def parse_etherchannel(etherchannel_out):
    """
    Parses 'show etherchannel summary'.

    C9500 IOS-XE format:
    ─────────────────────────────────────────────────────────────────────────
    Flags:  D - down        P - bundled in port-channel
            I - stand-alone s - suspended
            H - Hot-standby (LACP only)
            R - Layer3      S - Layer2
            U - in use      N - not in use, no aggregation
            f - failed to allocate aggregator

    Number of channel-groups in use: 1
    Number of aggregators:           1

    Group  Port-channel  Protocol    Ports
    ------+-------------+-----------+-----------------------------------------------
    1      Po1(SU)         LACP      Twe1/1/2(P)  Twe1/1/3(P)
    ─────────────────────────────────────────────────────────────────────────

    Returns:
      { intf_key_of_member: "Po1" }
      e.g. { "twe1/1/2": "Po1", "twe1/1/3": "Po1" }
    """
    result  = {}
    # Match the data rows — group number, port-channel, protocol, then member ports
    # Members look like:  Twe1/1/2(P)  Twe1/1/3(P)   flag in parentheses
    MEMBER_PAT = re.compile(r"([A-Za-z][\w/.-]+)\([A-Za-z]+\)")
    PO_PAT     = re.compile(r"(Po\d+)\(", re.I)

    for line in etherchannel_out.splitlines():
        line = line.strip()
        if not line or line.startswith(("Flags", "Number", "Group", "---", "R", "S")):
            continue
        # Find the Port-channel on this line
        po_m = PO_PAT.search(line)
        if not po_m:
            continue
        po_name = po_m.group(1)   # e.g. "Po1"
        # Find all member ports on the same line
        members = MEMBER_PAT.findall(line)
        for member in members:
            # Skip the port-channel itself appearing as a member token
            if member.lower().startswith("po"):
                continue
            result[intf_key(member)] = po_name

    return result


# ============================================================================
# ROW ASSEMBLY
# ============================================================================
def _sort_key(name):
    nums   = re.findall(r"\d+", name)
    prefix = re.sub(r"[\d/]", "", name).lower()
    return (prefix, [int(n) for n in nums])


def assemble_rows(up_map, raw_cmds):
    """
    Merges all parsed sources.
    up_map: {intf_key: display_name}
    Returns [OrderedDict per interface]
    """
    status_map  = parse_intf_status(raw_cmds["intf_status"])
    trunk_map   = parse_trunk_ports(raw_cmds["intf_trunk"])
    counter_map = parse_intf_full(raw_cmds["intf_full"], set(up_map.keys()))
    err_map     = parse_intf_counters_errors(raw_cmds["intf_errs"])
    sfp_map     = parse_transceiver(raw_cmds.get("transceiver", ""))
    lldp_map      = parse_lldp(raw_cmds.get("lldp", ""))
    mac_tbl_map   = parse_mac_table(raw_cmds.get("mac_table", ""))
    xcvr_dtl_map  = parse_transceiver_detail(raw_cmds.get("xcvr_detail", ""))
    po_member_map = parse_etherchannel(raw_cmds.get("etherchannel", ""))

    rows = []
    for key in sorted(up_map.keys(), key=lambda k: _sort_key(up_map[k])):
        display = up_map[key]
        row = OrderedDict()
        row["Interface"]   = display

        st   = status_map.get(key,  {})
        ctr  = counter_map.get(key, {})
        err  = err_map.get(key,     {})
        trk  = trunk_map.get(key,   {})
        sfp  = sfp_map.get(key,     {})
        lldp     = lldp_map.get(key,      {})
        mac_tbl  = mac_tbl_map.get(key,  [])
        xcvr_dtl = xcvr_dtl_map.get(key, {})

        row["Description"] = st.get("description") or ctr.get("Description", "")

        # ── Mode detection — priority order ──────────────────────────────────
        # 1. Explicit port-channel interface (Po1)
        # 2. Port-channel member (bundled into EtherChannel) — labeled as
        #    "pc-member (PoX)" so you know which bundle it belongs to
        # 3. In show interfaces trunk → trunk
        # 4. show interfaces status vlan field == "trunk" → trunk
        #    (catches trunk members NOT in 'show int trunk' like EtherChannel members
        #     and ports with no cdp / nonegotiate where IOS omits them)
        # 5. Default → access
        _po_parent  = po_member_map.get(key, "")
        _status_vlan = st.get("vlan", "").lower()

        if display.lower().startswith("port-channel") or key.startswith("po"):
            row["Mode"] = "port-channel"
        elif _po_parent:
            row["Mode"] = f"pc-member ({_po_parent})"
        elif key in trunk_map:
            row["Mode"] = "trunk"
        elif _status_vlan == "trunk":
            row["Mode"] = "trunk"
        else:
            row["Mode"] = "access"

        # Speed / Duplex — prefer show int status (negotiated values shown there)
        row["Speed"]  = st.get("speed")  or ctr.get("Speed",  "")
        row["Duplex"] = st.get("duplex") or ctr.get("Duplex", "")

        # VLAN info — trunk_map or fallback to status vlan field
        _is_trunk = row["Mode"] in ("trunk",) or row["Mode"].startswith("pc-member")
        if key in trunk_map:
            row["Native VLAN"]   = trk.get("native_vlan",  "")
            row["Allowed VLANs"] = trk.get("allowed_vlans", "")
        elif _is_trunk and _status_vlan == "trunk":
            # Member port not in trunk table — pull allowed VLANs from
            # the parent port-channel entry in trunk_map if available
            _po_key = intf_key(_po_parent) if _po_parent else ""
            _po_trk = trunk_map.get(_po_key, {})
            row["Native VLAN"]   = _po_trk.get("native_vlan",  "")
            row["Allowed VLANs"] = _po_trk.get("allowed_vlans", "")
        else:
            row["Native VLAN"]   = ""
            row["Allowed VLANs"] = st.get("vlan", "")

        # Counters from show interfaces
        for col in [
            "In Rate (bps)", "In Rate (pps)", "Out Rate (bps)", "Out Rate (pps)",
            "Pkts In", "Bytes In", "No Buffer",
            "Input Errors", "CRC", "Frame", "Overruns", "Ignored",
            "Runts", "Giants", "Throttles", "Dribble",
            "Pkts Out", "Bytes Out",
            "Output Errors", "Collisions", "Intf Resets",
            "Unknown Proto Drops", "Out Buffer Failures", "Out Bufs Swapped",
            "Late Collision", "Deferred", "Lost Carrier", "No Carrier",
        ]:
            row[col] = ctr.get(col, 0)

        # Counters from show interfaces counters errors
        for col in [
            "Align Err", "FCS Err", "Xmit Err", "Rcv Err",
            "Undersize", "Out Discards",
            "Single Col", "Multi Col", "Excess Col", "Carri Sen", "Oversize",
        ]:
            row[col] = err.get(col, 0)

        # SFP data from show interfaces transceiver
        row["SFP Type"]      = st.get("port_type", "")
        row["Temp (C)"]      = sfp.get("Temp (C)",        "N/A")
        row["Voltage (V)"]   = sfp.get("Voltage (V)",     "N/A")
        row["Current (mA)"]  = sfp.get("Current (mA)",    "N/A")
        row["Tx Power (dBm)"]= sfp.get("Tx Power (dBm)",  "N/A")
        row["Rx Power (dBm)"]= sfp.get("Rx Power (dBm)",  "N/A")
        row["SFP Alarms"]    = sfp.get("SFP Alarms",       "")
        row["SFP Status"]    = sfp.get("SFP Status",       "N/A")

        # Transceiver detail — thresholds and margin (fiber interfaces only)
        for _tc in ["Rx Hi Alarm (dBm)", "Rx Hi Warn (dBm)",
                    "Rx Lo Warn (dBm)",  "Rx Lo Alarm (dBm)",
                    "Tx Hi Alarm (dBm)", "Tx Hi Warn (dBm)",
                    "Tx Lo Warn (dBm)",  "Tx Lo Alarm (dBm)"]:
            row[_tc] = xcvr_dtl.get(_tc, "N/A")

        # Rx Margin = current Rx power minus Low Warn threshold
        # Negative margin means the signal is already below the warning threshold
        _rx_now     = sfp.get("Rx Power (dBm)", None)
        _rx_lo_warn = xcvr_dtl.get("Rx Lo Warn (dBm)", None)
        if isinstance(_rx_now, (int, float)) and isinstance(_rx_lo_warn, (int, float)):
            _margin = round(float(_rx_now) - float(_rx_lo_warn), 2)
            row["Rx Margin (dB)"] = _margin
            if _margin < 0:
                row["Fiber Rx Health"] = "BELOW WARN"
            elif _margin < 2:
                row["Fiber Rx Health"] = "NEAR WARN"
            else:
                row["Fiber Rx Health"] = "OK"
        elif xcvr_dtl:
            # Thresholds present but current Rx reading unavailable
            row["Rx Margin (dB)"]  = "N/A"
            row["Fiber Rx Health"] = "N/A"
        else:
            # Not a fiber/DOM interface
            row["Rx Margin (dB)"]  = ""
            row["Fiber Rx Health"] = ""

        # Also compare against Low Alarm for deeper diagnosis
        _rx_lo_alarm = xcvr_dtl.get("Rx Lo Alarm (dBm)", None)
        if isinstance(_rx_now, (int, float)) and isinstance(_rx_lo_alarm, (int, float)):
            if float(_rx_now) < float(_rx_lo_alarm):
                row["Fiber Rx Health"] = "BELOW ALARM"

        # LLDP neighbor info
        row["LLDP Neighbor"]     = lldp.get("hostname", "")
        row["LLDP Neighbor MAC"] = lldp.get("mac",      "")
        row["LLDP Neighbor IP"]  = lldp.get("ip",       "")

        # MAC address-table entries — always populate with every unique MAC
        # seen on this port regardless of whether LLDP found a neighbor.
        # Multiple MACs on one port = bridged device (camera with built-in switch,
        # unmanaged hub, IP phone + PC passthrough, etc.)
        row["MAC Table Entries"] = "; ".join(mac_tbl) if mac_tbl else ""

        row["Has Drops?"] = "YES" if any(
            _int(row.get(c, 0)) > 0 for c in DROP_COLUMNS
        ) else "NO"

        rows.append(row)
    return rows


# ============================================================================
# EXCEL WRITER
# ============================================================================
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
TRUNK_FILL = PatternFill("solid", fgColor="D6E4F0")
PC_FILL    = PatternFill("solid", fgColor="E8D5F5")
DROP_FILL  = PatternFill("solid", fgColor="FFCCCC")
DROP_YES   = PatternFill("solid", fgColor="C00000")
SFP_WARN   = PatternFill("solid", fgColor="FFE699")   # amber — SFP warning threshold
SFP_ALARM  = PatternFill("solid", fgColor="FF7F00")   # orange — SFP alarm threshold
SFP_ABSENT = PatternFill("solid", fgColor="BFBFBF")   # grey — NOT PRESENT
RX_OK      = PatternFill("solid", fgColor="C6EFCE")   # green  — Rx healthy
RX_NEAR    = PatternFill("solid", fgColor="FFEB9C")   # yellow — within 2 dB of threshold
RX_WARN    = PatternFill("solid", fgColor="FF7F00")   # orange — below Low Warn
RX_ALARM   = PatternFill("solid", fgColor="C00000")   # red    — below Low Alarm
ALT_FILL   = PatternFill("solid", fgColor="F5F5F5")
_THIN      = Side(style="thin", color="CCCCCC")
_BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr(cell):
    cell.fill      = HDR_FILL
    cell.font      = Font(color="FFFFFF", bold=True, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _BORDER


def write_excel(all_device_rows, out_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Per-device sheets ────────────────────────────────────────────────────
    used_sheet_names = set()
    for device, rows in all_device_rows.items():
        raw_name   = re.sub(r"[\\/*?:\[\]]", "_", device)[:31]
        sheet_name = raw_name
        suffix     = 2
        while sheet_name in used_sheet_names:
            # Truncate further to make room for the numeric suffix
            sheet_name = f"{raw_name[:28]}_{suffix}"
            suffix    += 1
        used_sheet_names.add(sheet_name)
        ws = wb.create_sheet(title=sheet_name)

        # ── Row 1: full device identity (never truncated) ────────────────────
        # Extract hostname and IP from the device label "HOSTNAME (IP)"
        _ip_m   = re.search(r"\(([^)]+)\)$", device)
        _ip     = _ip_m.group(1) if _ip_m else ""
        _host   = device.replace(f" ({_ip})", "").strip() if _ip else device.strip()
        title_cell = ws.cell(row=1, column=1,
                             value=f"Device: {_host}   IP: {_ip}   "
                                   f"Collected: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        title_cell.font      = Font(bold=True, size=11, color="FFFFFF")
        title_cell.fill      = PatternFill("solid", fgColor="1F4E79")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1,   end_column=len(COLUMNS))
        ws.row_dimensions[1].height = 22

        # ── Row 2: column headers ─────────────────────────────────────────────
        for ci, col in enumerate(COLUMNS, 1):
            _hdr(ws.cell(row=2, column=ci, value=col))
        ws.freeze_panes = "B3"          # freeze title + header, first col
        ws.row_dimensions[2].height = 36

        for ri, row_data in enumerate(rows, 3):
            mode   = row_data.get("Mode", "").lower()
            is_alt = (ri % 2 == 0)

            for ci, col in enumerate(COLUMNS, 1):
                val  = _safe_val(row_data.get(col, ""))
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border    = _BORDER
                cell.font      = Font(size=9)
                cell.alignment = Alignment(
                    horizontal="left" if ci == 1 else "center",
                    vertical="center")

                if mode == "trunk":
                    cell.fill = TRUNK_FILL
                elif mode == "port-channel":
                    cell.fill = PC_FILL
                elif is_alt:
                    cell.fill = ALT_FILL

                if col in DROP_COLUMNS and _int(val) > 0:
                    cell.fill = DROP_FILL
                    cell.font = Font(size=9, bold=True, color="8B0000")

                # SFP Status column colouring
                if col == "SFP Status":
                    if val == "ALARM":
                        cell.fill = SFP_ALARM
                        cell.font = Font(size=9, bold=True, color="FFFFFF")
                    elif val == "WARN":
                        cell.fill = SFP_WARN
                        cell.font = Font(size=9, bold=True, color="5C3D00")
                    elif val == "NOT PRESENT":
                        cell.fill = SFP_ABSENT
                        cell.font = Font(size=9, italic=True, color="404040")

                # SFP Alarms column — highlight any non-empty alarm string
                if col == "SFP Alarms" and val:
                    cell.fill = SFP_ALARM
                    cell.font = Font(size=9, bold=True, color="FFFFFF")

                if col == "Has Drops?" and val == "YES":
                    cell.fill = DROP_YES
                    cell.font = Font(size=9, bold=True, color="FFFFFF")

                # Fiber Rx Health — color-coded signal margin status
                if col == "Fiber Rx Health":
                    if val == "OK":
                        cell.fill = RX_OK
                        cell.font = Font(size=9, bold=True, color="1A6130")
                    elif val == "NEAR WARN":
                        cell.fill = RX_NEAR
                        cell.font = Font(size=9, bold=True, color="5C4000")
                    elif val == "BELOW WARN":
                        cell.fill = RX_WARN
                        cell.font = Font(size=9, bold=True, color="FFFFFF")
                    elif val == "BELOW ALARM":
                        cell.fill = RX_ALARM
                        cell.font = Font(size=9, bold=True, color="FFFFFF")

                # Rx Margin — highlight negative values (signal below threshold)
                if col == "Rx Margin (dB)" and val not in ("", "N/A"):
                    try:
                        if float(val) < 0:
                            cell.fill = RX_ALARM
                            cell.font = Font(size=9, bold=True, color="FFFFFF")
                        elif float(val) < 2:
                            cell.fill = RX_NEAR
                            cell.font = Font(size=9, bold=True, color="5C4000")
                    except (TypeError, ValueError):
                        pass

        # Auto-size columns
        for ci, col in enumerate(COLUMNS, 1):
            mx = len(col)
            for ri2 in range(3, ws.max_row + 1):   # skip title row
                v = ws.cell(row=ri2, column=ci).value
                if v:
                    mx = max(mx, len(str(v)))
            ws.column_dimensions[get_column_letter(ci)].width = max(8, min(40, mx + 2))

        # Summary footer
        sr    = ws.max_row + 2
        total = len(rows)
        trunks = sum(1 for r in rows if r.get("Mode") == "trunk")
        drops  = sum(1 for r in rows if r.get("Has Drops?") == "YES")
        footer_cell = ws.cell(row=sr, column=1,
                value=_safe_val(f"Total interfaces: {total}  |  Trunks: {trunks}  |  "
                                f"Access/Other: {total - trunks}  |  Drop ports: {drops}"))
        footer_cell.font = Font(bold=True, size=9, color="1F4E79")

        logger.info(f"  Sheet '{sheet_name}' [{_host}]: {total} interfaces, {drops} with drops")

    # ── Run Summary sheet (index 0) ──────────────────────────────────────────
    ws_s = wb.create_sheet(title="Run Summary", index=0)
    ws_s.freeze_panes = "A2"
    hdrs = ["Device", "IP", "Total Intfs", "Trunks", "Access", "Drop Ports"]
    widths = [32, 18, 12, 10, 10, 12]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        _hdr(ws_s.cell(row=1, column=ci, value=h))
        ws_s.column_dimensions[get_column_letter(ci)].width = w
    ws_s.row_dimensions[1].height = 28

    for ri, (label, rows) in enumerate(all_device_rows.items(), 2):
        ip_m  = re.search(r"\(([^)]+)\)", label)
        ip    = ip_m.group(1) if ip_m else ""
        dev   = label.replace(f" ({ip})", "") if ip else label
        total  = len(rows)
        trunks = sum(1 for r in rows if r.get("Mode") == "trunk")
        drops  = sum(1 for r in rows if r.get("Has Drops?") == "YES")

        ws_s.cell(row=ri, column=1, value=dev)
        ws_s.cell(row=ri, column=2, value=ip)
        ws_s.cell(row=ri, column=3, value=total)
        ws_s.cell(row=ri, column=4, value=trunks)
        ws_s.cell(row=ri, column=5, value=total - trunks)
        dc = ws_s.cell(row=ri, column=6, value=drops)
        if drops > 0:
            dc.fill = DROP_FILL
            dc.font = Font(bold=True, color="8B0000", size=9)

    wb.save(out_path)
    logger.info(f"Excel saved -> {out_path}")


# ============================================================================
# MAIN
# ============================================================================
def load_targets(path):
    if not os.path.exists(path):
        print(f"[ERROR] Target file not found: {path}")
        sys.exit(1)
    with open(path) as f:
        return [ln.strip() for ln in f
                if ln.strip() and not ln.startswith("#")]


def main():
    global logger, AGG_SWITCH_IP

    ap = argparse.ArgumentParser(description="Cisco Port Drop Auditor")
    ap.add_argument("--targets",  required=True)
    ap.add_argument("--agg",      default=AGG_SWITCH_IP)
    ap.add_argument("--no-jump",  action="store_true")
    ap.add_argument("--out",      default="")
    args = ap.parse_args()
    AGG_SWITCH_IP = args.agg

    logger, log_file = setup_logging()
    logger.info("=" * 60)
    logger.info("Cisco Port Drop Auditor")
    logger.info("=" * 60)

    targets = load_targets(args.targets)
    logger.info(f"Loaded {len(targets)} target(s) from {args.targets}")

    if not args.no_jump:
        if not connect_to_agg(AGG_SWITCH_IP):
            logger.error("Cannot reach aggregate. Exiting.")
            sys.exit(1)

    all_device_rows = {}

    for target_ip in targets:
        logger.info(f"\n-> {target_ip}")

        collected = False
        for attempt in range(1, RETRY_COUNT + 2):
            if attempt > 1:
                logger.warning(f"  Socket dropped mid-collection — waiting {RETRY_DELAY}s then retrying {target_ip} (attempt {attempt}/{RETRY_COUNT + 1}) ...")
                time.sleep(RETRY_DELAY)
                # Re-establish the agg connection before jumping again
                if not args.no_jump:
                    if not connect_to_agg(AGG_SWITCH_IP):
                        logger.error("  Cannot reconnect to aggregate — giving up on this device.")
                        break

            # ── Connect / jump ──────────────────────────────────────────────
            if args.no_jump:
                _, shell, hostname = connect_direct(target_ip)
                if not hostname:
                    logger.error(f"  Direct connect failed: {target_ip}")
                    break  # credential failure — retrying won't help
            else:
                hostname = jump_to_target(target_ip)
                if not hostname:
                    break  # credential failure — retrying won't help
                shell = agg_shell

            # ── Collect commands ────────────────────────────────────────────
            try:
                raw = run_commands(shell)
                collected = True
            except OSError as e:
                logger.warning(f"  Socket closed during collection ({e})")
                continue  # retry the whole login + collect cycle

            # ── Parse & store ───────────────────────────────────────────────
            up_map = parse_up_interfaces(raw["ip_brief"])
            logger.info(f"  {len(up_map)} up/up interfaces found")

            if up_map:
                rows = assemble_rows(up_map, raw)
                all_device_rows[f"{hostname} ({target_ip})"] = rows
            else:
                logger.warning("  No up/up interfaces - skipping sheet.")

            break  # success — move to next device

        if not collected:
            logger.error(f"  Permanently failed to collect data from {target_ip} — skipping.")

        if not args.no_jump:
            return_to_agg()

    if all_device_rows:
        ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = args.out or f"port_audit_{ts}.xlsx"
        write_excel(all_device_rows, out)
        logger.info(f"\nOutput -> {out}")
    else:
        logger.warning("No data collected - Excel not written.")

    stats.print_summary()
    logger.info(f"Log -> {log_file}")


if __name__ == "__main__":
    main()
